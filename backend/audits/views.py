"""Views for the audits application."""
from __future__ import annotations

import csv
import io
import json
import mimetypes
import os
from collections.abc import Iterable
from datetime import date, datetime, time, timedelta
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.signing import BadSignature, SignatureExpired
from django.db.models import (
    Case,
    Count,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    QuerySet,
    Subquery,
    Value,
    When,
)
from django.db.models.functions import Cast
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import ListView, TemplateView

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from accounts.models import UserProfile
from accounts.permissions import (
    RoleQuerysetMixin,
    RoleRequiredMixin,
    is_admin,
    restrict_queryset_for_user,
)

from .forms import AuditRequestChangesForm
from .models import (
    AttachmentLimits,
    Audit,
    AuditAttachment,
    AuditLogEntry,
    AuditResponse,
    AuditSignature,
    OfflineSyncBatch,
)
from .tokens import read_attachment_token
from .reporting import build_audit_report
from .services import build_catalog_snapshot_for_user, build_checklist_structure


class AuditListView(RoleQuerysetMixin, ListView):
    """List of audits available to the current user with filtering support."""

    model = Audit
    template_name = "audits/audit_list.html"
    context_object_name = "audits"
    paginate_by = 20
    ordering = "-created_at"
    allowed_roles = (UserProfile.Roles.AUDITOR, UserProfile.Roles.ADMIN)

    search_param = "q"
    status_param = "status"
    period_param = "period"
    review_param = "review"

    _role_filtered_queryset: QuerySet[Audit] | None = None
    _status_filter: str | None = None
    _period_filter_value: str | None = None
    _search_query: str | None = None
    _review_filter_value: str | None = None

    def dispatch(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Audit]:  # type: ignore[override]
        queryset = (
            super()
            .get_queryset()
            .select_related("elevator", "elevator__building", "created_by__profile")
        )
        self._role_filtered_queryset = queryset
        queryset = self.filter_by_status(queryset)
        queryset = self.filter_by_period(queryset)
        queryset = self.filter_by_review_state(queryset)
        queryset = self.apply_search(queryset)
        return queryset.order_by(self.ordering)

    # --- filters ---------------------------------------------------------

    def get_status_choices(self) -> list[tuple[str, str]]:
        return [("", _("Все статусы"))] + list(Audit.Status.choices)

    def get_period_choices(self) -> list[tuple[str, str]]:
        return [
            ("", _("За всё время")),
            ("7", _("За последние 7 дней")),
            ("30", _("За последние 30 дней")),
            ("90", _("За последние 90 дней")),
        ]

    def get_review_choices(self) -> list[tuple[str, str]]:
        return [
            ("", _("Все аудиты")),
            ("pending", _("Ожидают проверки")),
            ("active", _("В работе")),
            ("reviewed", _("Просмотренные")),
        ]

    def get_status_filter(self) -> str:
        if self._status_filter is not None:
            return self._status_filter
        raw = self.request.GET.get(self.status_param, "").strip()
        valid_values = {value for value, _ in self.get_status_choices() if value}
        self._status_filter = raw if raw in valid_values else ""
        return self._status_filter

    def get_period_filter(self) -> int | None:
        raw = self.request.GET.get(self.period_param, "").strip()
        self._period_filter_value = raw
        if not raw:
            return None
        valid_values = {value for value, _ in self.get_period_choices() if value}
        if raw not in valid_values:
            return None
        try:
            return int(raw)
        except ValueError:
            return None

    def get_search_query(self) -> str:
        if self._search_query is None:
            self._search_query = self.request.GET.get(self.search_param, "").strip()
        return self._search_query

    def get_review_filter(self) -> str:
        if not is_admin(self.request.user):
            return ""
        if self._review_filter_value is not None:
            return self._review_filter_value
        raw = self.request.GET.get(self.review_param, "").strip()
        valid_values = {value for value, _ in self.get_review_choices() if value}
        self._review_filter_value = raw if raw in valid_values else ""
        return self._review_filter_value

    def filter_by_status(self, queryset: QuerySet[Audit]) -> QuerySet[Audit]:
        status = self.get_status_filter()
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def filter_by_period(self, queryset: QuerySet[Audit]) -> QuerySet[Audit]:
        period_days = self.get_period_filter()
        if period_days:
            threshold = timezone.now() - timedelta(days=period_days)
            queryset = queryset.filter(created_at__gte=threshold)
        return queryset

    def filter_by_review_state(self, queryset: QuerySet[Audit]) -> QuerySet[Audit]:
        if not is_admin(self.request.user):
            return queryset
        review_value = self.get_review_filter()
        if review_value == "pending":
            return queryset.filter(status=Audit.Status.SUBMITTED)
        if review_value == "active":
            return queryset.filter(status__in=[Audit.Status.DRAFT, Audit.Status.IN_PROGRESS])
        if review_value == "reviewed":
            return queryset.filter(status=Audit.Status.REVIEWED)
        return queryset

    def apply_search(self, queryset: QuerySet[Audit]) -> QuerySet[Audit]:
        query = self.get_search_query()
        if not query:
            return queryset
        return queryset.filter(
            Q(elevator__identifier__icontains=query)
            | Q(elevator__building__address__icontains=query)
        )

    def get_preserved_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        non_empty = {key: value for key, value in params.items() if value}
        return urlencode(non_empty)

    # --- context ---------------------------------------------------------

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        base_queryset = self._role_filtered_queryset or self.model.objects.none()
        summary = {
            entry["status"]: entry["total"]
            for entry in base_queryset.values("status").annotate(total=Count("id"))
        }
        total_count = base_queryset.count()

        status_filter = self.get_status_filter()
        status_filters: list[dict[str, object]] = []
        for value, label in self.get_status_choices():
            if value:
                count = summary.get(value, 0)
            else:
                count = total_count
            status_filters.append(
                {
                    "value": value,
                    "label": label,
                    "count": count,
                    "selected": value == status_filter,
                }
            )

        period_value = self._period_filter_value or ""
        period_filters = [
            {
                "value": value,
                "label": label,
                "selected": value == period_value,
            }
            for value, label in self.get_period_choices()
        ]

        active_filters: list[dict[str, str]] = []
        if status_filter:
            status_label = next(
                (label for value, label in self.get_status_choices() if value == status_filter),
                status_filter,
            )
            active_filters.append({"label": _("Статус"), "value": status_label})
        if period_value:
            period_label = next(
                (label for value, label in self.get_period_choices() if value == period_value),
                period_value,
            )
            active_filters.append({"label": _("Период"), "value": period_label})
        search_query = self.get_search_query()
        if search_query:
            active_filters.append({"label": _("Поиск"), "value": search_query})

        review_value = self.get_review_filter()
        review_filters: list[dict[str, object]] = []
        if is_admin(self.request.user):
            for value, label in self.get_review_choices():
                review_filters.append(
                    {
                        "value": value,
                        "label": label,
                        "selected": value == review_value,
                    }
                )
            if review_value:
                review_label = next(
                    (label for value, label in self.get_review_choices() if value == review_value),
                    review_value,
                )
                active_filters.append({"label": _("Проверка"), "value": review_label})

        context.update(
            {
                "status_filters": status_filters,
                "period_filters": period_filters,
                "active_filters": active_filters,
                "search_query": search_query,
                "search_param": self.search_param,
                "status_param": self.status_param,
                "period_param": self.period_param,
                "review_filters": review_filters,
                "review_param": self.review_param,
                "querystring": self.get_preserved_querystring(),
                "AuditStatus": Audit.Status,
                "catalog_snapshot_url": reverse("catalog-snapshot"),
                "is_admin": is_admin(self.request.user),
            }
        )
        badge_classes = self.get_status_badge_classes()
        default_class = "border-border-subtle bg-surface-subtle text-ink-600"
        page_obj = context.get("page_obj")
        if page_obj is None:
            possible_page = context.get(self.context_object_name)
            if hasattr(possible_page, "object_list"):
                page_obj = possible_page
        if hasattr(page_obj, "object_list"):
            for audit in getattr(page_obj, "object_list", []):
                setattr(audit, "badge_class", badge_classes.get(audit.status, default_class))
        context["default_status_badge_class"] = default_class
        return context

    @staticmethod
    def get_status_badge_classes() -> dict[str, str]:
        return {
            Audit.Status.DRAFT: "border-border bg-surface-muted text-ink-700",
            Audit.Status.IN_PROGRESS: "border-warning-200 bg-warning-50 text-warning-700",
            Audit.Status.SUBMITTED: "border-brand-200 bg-brand-50 text-brand-700",
            Audit.Status.REVIEWED: "border-success-200 bg-success-50 text-success-700",
        }


class AttachmentDownloadView(LoginRequiredMixin, View):
    """Serve audit attachments through authenticated, signed URLs."""

    http_method_names = ["get"]

    def get(self, request, token: str, *args, **kwargs) -> HttpResponse:  # type: ignore[override]
        try:
            attachment_id = read_attachment_token(token)
        except SignatureExpired as exc:
            raise Http404("Ссылка для скачивания истекла.") from exc
        except BadSignature as exc:
            raise Http404("Недействительная ссылка для скачивания.") from exc

        attachment = (
            AuditAttachment.objects.select_related("response__audit__created_by")
            .filter(pk=attachment_id)
            .first()
        )
        if attachment is None or not attachment.file:
            raise Http404("Вложение не найдено.")

        audit = attachment.response.audit
        user = request.user
        if not self._user_can_access(user, audit.created_by_id):
            raise Http404("Вложение не найдено.")

        storage = attachment.file.storage
        if not storage.exists(attachment.file.name):
            raise Http404("Файл вложения отсутствует на сервере.")

        filename = os.path.basename(attachment.file.name)
        content_type, _ = mimetypes.guess_type(filename)
        file_handle = attachment.file.open("rb")

        response = FileResponse(file_handle, as_attachment=True, filename=filename)
        if content_type:
            response.headers["Content-Type"] = content_type
        return response

    @staticmethod
    def _user_can_access(user: object, author_id: int | None) -> bool:
        if not hasattr(user, "is_authenticated") or not user.is_authenticated:
            return False
        if is_admin(user):
            return True
        return getattr(user, "pk", None) == author_id


class AuditExportBaseView(RoleRequiredMixin):
    """Shared helpers for audit export views."""

    allowed_roles = (UserProfile.Roles.AUDITOR, UserProfile.Roles.ADMIN)

    _audit_instance: Audit | None = None
    _report_cache: dict[str, object] | None = None
    _checklist_structure: dict[str, object] | None = None

    def dispatch(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Audit]:  # type: ignore[override]
        response_qs = (
            AuditResponse.objects.select_related(
                "question",
                "question__section",
                "question__section__category",
            )
            .prefetch_related(
                Prefetch(
                    "attachments",
                    queryset=AuditAttachment.objects.order_by("uploaded_at"),
                )
            )
            .order_by(
                "question__section__category__order",
                "question__section__order",
                "question__order",
                "question_id",
            )
        )
        base_qs = (
            Audit.objects.select_related("elevator", "elevator__building", "created_by")
            .prefetch_related(Prefetch("responses", queryset=response_qs))
            .order_by("-created_at", "-id")
        )
        return restrict_queryset_for_user(
            base_qs, self.request.user, auditor_field="created_by"
        )

    def get_audit(self) -> Audit:
        if self._audit_instance is None:
            pk = self.kwargs.get("pk")
            audit = self.get_queryset().filter(pk=pk).first()
            if audit is None:
                raise Http404("Аудит не найден или недоступен.")
            self._audit_instance = audit
        return self._audit_instance

    def get_checklist_structure(self) -> dict[str, object]:
        if self._checklist_structure is None:
            self._checklist_structure = build_checklist_structure()
        return self._checklist_structure

    def get_report(self) -> dict[str, object]:
        if self._report_cache is None:
            self._report_cache = build_audit_report(
                self.get_audit(), checklist_structure=self.get_checklist_structure()
            )
        return self._report_cache

    def format_datetime(self, value) -> str:
        if value is None:
            return ""
        return timezone.localtime(value).strftime("%d.%m.%Y %H:%M")

    def format_flag(self, value: bool) -> str:
        return str(_("Да")) if value else str(_("Нет"))

    def format_attachments(self, attachments: list[AuditAttachment]) -> str:
        if not attachments:
            return ""
        items: list[str] = []
        for attachment in attachments:
            url = attachment.get_download_url()
            if hasattr(self, "request"):
                url = self.request.build_absolute_uri(url)
            caption = (attachment.caption or "").strip()
            if caption:
                items.append(f"{caption} ({url})")
            else:
                items.append(url)
        return "; ".join(items)

    def get_question_answer(self, question: dict[str, object]) -> str:
        if question.get("type") == "text":
            return str(question.get("answer_display") or "—")
        return str(question.get("value_display") or "—")

    def get_question_comment(self, question: dict[str, object]) -> str:
        if question.get("type") == "text":
            return ""
        return str(question.get("comment_display") or "")


class AuditDetailView(AuditExportBaseView, TemplateView):
    """Detailed view for administrators to review a specific audit."""

    template_name = "audits/admin_audit_detail.html"
    allowed_roles = (UserProfile.Roles.ADMIN,)

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        audit = self.get_audit()
        report = self.get_report()
        badge_classes = AuditListView.get_status_badge_classes()
        default_badge = "border-border-subtle bg-surface-subtle text-ink-600"
        request_changes_form = kwargs.get("request_changes_form") or context.get(
            "request_changes_form"
        )
        if request_changes_form is None:
            request_changes_form = AuditRequestChangesForm()

        context.update(
            {
                "audit": audit,
                "report": report,
                "summary": report.get("summary", {}),
                "checklist": report.get("checklist", []),
                "object_info": report.get("object_info", []),
                "object_info_has_values": report.get("object_info_has_values", False),
                "object_info_has_extra": report.get("object_info_has_extra", False),
                "can_mark_reviewed": audit.status == Audit.Status.SUBMITTED,
                "can_request_changes": audit.status == Audit.Status.SUBMITTED,
                "request_changes_form": request_changes_form,
                "back_url": self.get_back_url(),
                "status_badge_class": badge_classes.get(audit.status, default_badge),
                "default_status_badge_class": default_badge,
                "AuditStatus": Audit.Status,
            }
        )
        return context

    def get_back_url(self) -> str:
        candidate = self.request.GET.get("next") or self.request.GET.get("back")
        if candidate and url_has_allowed_host_and_scheme(
            candidate,
            allowed_hosts={self.request.get_host()},
            require_https=self.request.is_secure(),
        ):
            return candidate
        return reverse("audits:audit-list")


class AuditMarkReviewedView(RoleRequiredMixin, View):
    """Mark an audit as reviewed from the administrator interface."""

    allowed_roles = (UserProfile.Roles.ADMIN,)

    def get_queryset(self) -> QuerySet[Audit]:
        base_qs = Audit.objects.select_related(
            "elevator",
            "elevator__building",
            "created_by",
            "created_by__profile",
        )
        return restrict_queryset_for_user(base_qs, self.request.user, auditor_field="created_by")

    def post(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        pk = kwargs.get("pk")
        audit = self.get_queryset().filter(pk=pk).first()
        if audit is None:
            raise Http404("Аудит не найден или недоступен.")

        if audit.status == Audit.Status.REVIEWED:
            messages.info(request, _("Аудит уже отмечен как просмотренный."))
        elif audit.status != Audit.Status.SUBMITTED:
            messages.warning(request, _("Отметить просмотр можно только для отправленных аудитов."))
        else:
            audit.mark_reviewed(actor=request.user)
            messages.success(request, _("Аудит отмечен как просмотренный."))

        next_url = (request.POST.get("next") or "").strip()
        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            return redirect(next_url)
        return redirect("audits:audit-detail", pk=audit.pk)


class AuditRequestChangesView(AuditDetailView):
    """Handle administrator requests for auditor corrections."""

    http_method_names = ["post"]

    def post(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()

        audit = self.get_audit()
        form = AuditRequestChangesForm(request.POST)

        if audit.status != Audit.Status.SUBMITTED:
            messages.warning(
                request,
                _("Запросить правки можно только для отправленных аудитов."),
            )
            return redirect("audits:audit-detail", pk=audit.pk)

        if not form.is_valid():
            context = self.get_context_data(request_changes_form=form)
            return self.render_to_response(context, status=400)

        audit.request_changes(actor=request.user, message=form.cleaned_data["message"])
        messages.success(request, _("Запрос на правки отправлен автору аудита."))

        next_url = (request.POST.get("next") or "").strip()
        if next_url and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        ):
            return redirect(next_url)
        return redirect("audits:audit-detail", pk=audit.pk)


class OfflineObjectInfoView(RoleRequiredMixin, TemplateView):
    """Offline-friendly form for filling audit object information."""

    template_name = "audits/offline_object_info.html"
    allowed_roles = (UserProfile.Roles.AUDITOR, UserProfile.Roles.ADMIN)

    def dispatch(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        client_id = self.request.GET.get("client_id", "").strip()
        snapshot = build_catalog_snapshot_for_user(self.request.user)
        context.update(
            {
                "client_id": client_id,
                "catalog_snapshot_url": reverse("catalog-snapshot"),
                "object_info_fields": snapshot.get("object_fields", []),
                "catalog_payload": {
                    "buildings": snapshot.get("buildings", []),
                    "elevators": snapshot.get("elevators", []),
                },
                "catalog_metadata": {
                    "generated_at": snapshot.get("generated_at"),
                },
                "return_url": reverse("audits:audit-list"),
            }
        )
        return context


class OfflineChecklistView(RoleRequiredMixin, TemplateView):
    """Offline-friendly checklist form for auditors."""

    template_name = "audits/offline_checklist.html"
    allowed_roles = (UserProfile.Roles.AUDITOR, UserProfile.Roles.ADMIN)

    def dispatch(self, request, *args, **kwargs):  # type: ignore[override]
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        client_id = self.request.GET.get("client_id", "").strip()
        checklist = build_checklist_structure()
        limits = AttachmentLimits()

        context.update(
            {
                "client_id": client_id,
                "checklist": checklist,
                "attachment_limits": limits,
                "max_attachment_size_mb": limits.max_size_mb,
                "return_url": reverse("audits:audit-list"),
            }
        )
        return context


class AuditPrintView(AuditExportBaseView, TemplateView):
    """Render a printable HTML representation of an audit."""

    template_name = "audits/export_print.html"

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        report = self.get_report()
        checklist_display: list[dict[str, object]] = []
        for category in report["checklist"]:
            section_entries: list[dict[str, object]] = []
            for section in category["sections"]:
                question_entries: list[dict[str, object]] = []
                for question in section["questions"]:
                    attachments_info = [
                        {
                            "caption": attachment.caption,
                            "url": self.request.build_absolute_uri(
                                attachment.get_download_url()
                            ),
                        }
                        for attachment in question.get("attachments", [])
                    ]
                    question_copy = question.copy()
                    question_copy["attachments_for_display"] = attachments_info
                    question_entries.append(question_copy)
                section_entries.append(
                    {
                        "title": section.get("title"),
                        "description": section.get("description"),
                        "questions": question_entries,
                    }
                )
            checklist_display.append(
                {
                    "name": category.get("name"),
                    "sections": section_entries,
                }
            )
        context.update(
            {
                "audit": self.get_audit(),
                "report": report,
                "object_info": report["object_info"],
                "object_info_has_values": report["object_info_has_values"],
                "checklist": checklist_display,
                "summary": report["summary"],
            }
        )
        return context


class AuditCSVExportView(AuditExportBaseView, View):
    """Generate a CSV export for an audit."""

    http_method_names = ["get"]

    def get(self, request, *args, **kwargs):  # type: ignore[override]
        audit = self.get_audit()
        report = self.get_report()
        summary = report["summary"]

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)

        metadata_rows = [
            ("Аудит", f"#{audit.pk}"),
            ("Объект", str(audit.elevator.building)),
            ("Лифт", audit.elevator.identifier),
            ("Статус", audit.get_status_display()),
            ("Автор", str(audit.created_by)),
            ("Суммарный балл", audit.total_score),
            ("Плановая дата", audit.planned_date.isoformat() if audit.planned_date else ""),
            ("Начато", self.format_datetime(audit.started_at)),
            ("Завершено", self.format_datetime(audit.finished_at)),
            ("Всего вопросов", summary["total_questions"]),
            ("Заполнено", summary["answered_questions"]),
            ("Комментарии", summary["comments_total"]),
            ("Вложения", summary["attachments_total"]),
            ("Пометки", summary["flagged_total"]),
        ]
        writer.writerows(metadata_rows)
        writer.writerow([])
        writer.writerow(
            [
                "Категория",
                "Раздел",
                "Вопрос",
                "Ответ",
                "Комментарий",
                "Пометка",
                "Вложения",
            ]
        )

        for category in report["checklist"]:
            for section in category["sections"]:
                for question in section["questions"]:
                    attachments = self.format_attachments(question.get("attachments", []))
                    writer.writerow(
                        [
                            category.get("name", ""),
                            section.get("title", ""),
                            question.get("text", ""),
                            self.get_question_answer(question),
                            self.get_question_comment(question),
                            self.format_flag(bool(question.get("is_flagged"))),
                            attachments,
                        ]
                    )

        response = HttpResponse(
            buffer.getvalue().encode("utf-8-sig"),
            content_type="text/csv; charset=utf-8",
        )
        response["Content-Disposition"] = f'attachment; filename="audit-{audit.pk}.csv"'
        return response


class AuditExcelExportView(AuditExportBaseView, View):
    """Generate an XLSX spreadsheet with audit results."""

    http_method_names = ["get"]

    def get(self, request, *args, **kwargs):  # type: ignore[override]
        audit = self.get_audit()
        report = self.get_report()
        summary = report["summary"]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Аудит"

        workbook.properties.title = f"Audit #{audit.pk}"

        meta_rows = [
            ["Аудит", f"#{audit.pk}"],
            ["Объект", str(audit.elevator.building)],
            ["Лифт", audit.elevator.identifier],
            ["Статус", audit.get_status_display()],
            ["Автор", str(audit.created_by)],
            ["Суммарный балл", audit.total_score],
            ["Плановая дата", audit.planned_date.isoformat() if audit.planned_date else ""],
            ["Начато", self.format_datetime(audit.started_at)],
            ["Завершено", self.format_datetime(audit.finished_at)],
            ["Всего вопросов", summary["total_questions"]],
            ["Заполнено", summary["answered_questions"]],
            ["Комментарии", summary["comments_total"]],
            ["Вложения", summary["attachments_total"]],
            ["Пометки", summary["flagged_total"]],
        ]
        for row in meta_rows:
            sheet.append(row)

        sheet.append([])
        header_row_index = sheet.max_row + 1
        sheet.append(
            [
                "Категория",
                "Раздел",
                "Вопрос",
                "Ответ",
                "Комментарий",
                "Пометка",
                "Вложения",
            ]
        )

        for category in report["checklist"]:
            for section in category["sections"]:
                for question in section["questions"]:
                    attachments = self.format_attachments(question.get("attachments", []))
                    sheet.append(
                        [
                            category.get("name", ""),
                            section.get("title", ""),
                            question.get("text", ""),
                            self.get_question_answer(question),
                            self.get_question_comment(question),
                            self.format_flag(bool(question.get("is_flagged"))),
                            attachments,
                        ]
                    )

        column_widths = {
            1: 24,
            2: 24,
            3: 60,
            4: 18,
            5: 40,
            6: 12,
            7: 50,
        }
        for column_index, width in column_widths.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = width

        sheet.freeze_panes = f"A{header_row_index + 1}"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = f'attachment; filename="audit-{audit.pk}.xlsx"'
        return response


class AuditLogEntryListView(RoleRequiredMixin, ListView):
    """User-facing log viewer for administrators."""

    model = AuditLogEntry
    template_name = "audits/admin_log_list.html"
    context_object_name = "log_entries"
    paginate_by = 25
    ordering = "-created_at"
    allowed_roles = (UserProfile.Roles.ADMIN,)

    start_param = "start"
    end_param = "end"
    audit_param = "audit"
    entity_type_param = "entity_type"

    _start_date: date | None = None
    _end_date: date | None = None
    _audit_filter: int | None = None
    _entity_type_filter: str | None = None

    ENTITY_TYPE_LABELS: dict[str, str] = {
        "audits.audit": _("Аудит"),
        "audits.auditresponse": _("Ответ чек-листа"),
        "audits.auditattachment": _("Вложение"),
        "audits.auditsignature": _("Подпись"),
        "audits.offlinesyncbatch": _("Офлайн-пакет"),
    }

    def get(self, request, *args, **kwargs):  # type: ignore[override]
        if request.GET.get("export") == "csv":
            queryset = self.get_filtered_queryset()
            return self._export_csv(queryset)
        return super().get(request, *args, **kwargs)

    def get_queryset(self) -> QuerySet[AuditLogEntry]:  # type: ignore[override]
        return self.get_filtered_queryset()

    def get_filtered_queryset(self) -> QuerySet[AuditLogEntry]:
        queryset = (
            super()
            .get_queryset()
            .select_related("user")
            .annotate(**self._audit_annotations())
        )
        queryset = self._filter_by_dates(queryset)
        queryset = self._filter_by_entity_type(queryset)
        queryset = self._filter_by_audit(queryset)
        return queryset.order_by(self.ordering)

    def _audit_annotations(self) -> dict[str, Case]:
        audit_lookup = Subquery(
            Audit.objects.filter(pk=Cast(OuterRef("entity_id"), IntegerField()))
            .values("pk")[:1]
        )
        response_lookup = Subquery(
            AuditResponse.objects.filter(
                pk=Cast(OuterRef("entity_id"), IntegerField())
            ).values("audit_id")[:1]
        )
        attachment_lookup = Subquery(
            AuditAttachment.objects.filter(
                pk=Cast(OuterRef("entity_id"), IntegerField())
            ).values("response__audit_id")[:1]
        )
        signature_lookup = Subquery(
            AuditSignature.objects.filter(
                pk=Cast(OuterRef("entity_id"), IntegerField())
            ).values("audit_id")[:1]
        )

        return {
            "related_audit_id": Case(
                When(
                    entity_type="audits.audit",
                    entity_id__regex=r"^\d+$",
                    then=audit_lookup,
                ),
                When(
                    entity_type="audits.auditresponse",
                    entity_id__regex=r"^\d+$",
                    then=response_lookup,
                ),
                When(
                    entity_type="audits.auditattachment",
                    entity_id__regex=r"^\d+$",
                    then=attachment_lookup,
                ),
                When(
                    entity_type="audits.auditsignature",
                    entity_id__regex=r"^\d+$",
                    then=signature_lookup,
                ),
                default=Value(None),
                output_field=IntegerField(),
            )
        }

    def _filter_by_dates(self, queryset: QuerySet[AuditLogEntry]) -> QuerySet[AuditLogEntry]:
        start_date = self.get_start_date()
        if start_date:
            start_dt = timezone.make_aware(
                datetime.combine(start_date, time.min),
                timezone.get_current_timezone(),
            )
            queryset = queryset.filter(created_at__gte=start_dt)

        end_date = self.get_end_date()
        if end_date:
            exclusive_end = end_date + timedelta(days=1)
            end_dt = timezone.make_aware(
                datetime.combine(exclusive_end, time.min),
                timezone.get_current_timezone(),
            )
            queryset = queryset.filter(created_at__lt=end_dt)
        return queryset

    def _filter_by_entity_type(
        self, queryset: QuerySet[AuditLogEntry]
    ) -> QuerySet[AuditLogEntry]:
        entity_type = self.get_entity_type_filter()
        if entity_type and entity_type in self.ENTITY_TYPE_LABELS:
            queryset = queryset.filter(entity_type=entity_type)
        return queryset

    def _filter_by_audit(
        self, queryset: QuerySet[AuditLogEntry]
    ) -> QuerySet[AuditLogEntry]:
        audit_id = self.get_audit_filter()
        if audit_id is None:
            return queryset
        audit_id_str = str(audit_id)
        return queryset.filter(
            Q(entity_type="audits.audit", entity_id=audit_id_str)
            | Q(payload__audit_id=audit_id)
            | Q(payload__audit_id=audit_id_str)
            | Q(related_audit_id=audit_id)
        )

    def get_start_date(self) -> date | None:
        if self._start_date is not None:
            return self._start_date
        raw = self.request.GET.get(self.start_param, "").strip()
        self._start_date = parse_date(raw)
        return self._start_date

    def get_end_date(self) -> date | None:
        if self._end_date is not None:
            return self._end_date
        raw = self.request.GET.get(self.end_param, "").strip()
        self._end_date = parse_date(raw)
        return self._end_date

    def get_audit_filter(self) -> int | None:
        if self._audit_filter is not None:
            return self._audit_filter
        raw = self.request.GET.get(self.audit_param, "").strip()
        if not raw:
            self._audit_filter = None
            return None
        try:
            self._audit_filter = int(raw)
        except (TypeError, ValueError):
            self._audit_filter = None
        return self._audit_filter

    def get_entity_type_filter(self) -> str:
        if self._entity_type_filter is not None:
            return self._entity_type_filter
        raw = self.request.GET.get(self.entity_type_param, "").strip()
        if raw in self.ENTITY_TYPE_LABELS:
            self._entity_type_filter = raw
        else:
            self._entity_type_filter = ""
        return self._entity_type_filter

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        entries = self._collect_entries(context)
        audit_map = self._prepare_audit_map(entries)
        for entry in entries:
            audit_obj = audit_map.get(getattr(entry, "related_audit_id", None))
            if audit_obj is not None:
                setattr(entry, "related_audit", audit_obj)
            setattr(entry, "entity_label", self.ENTITY_TYPE_LABELS.get(entry.entity_type, entry.entity_type))
            setattr(entry, "payload_pretty", self._format_payload(entry.payload))

        context.update(
            {
                "start_value": self._format_date(self.get_start_date()),
                "end_value": self._format_date(self.get_end_date()),
                "audit_value": self.get_audit_filter() or "",
                "entity_type_choices": self._build_entity_type_choices(),
                "entity_type_value": self.get_entity_type_filter(),
                "querystring": self.get_preserved_querystring(),
                "start_param": self.start_param,
                "end_param": self.end_param,
                "audit_param": self.audit_param,
                "entity_type_param": self.entity_type_param,
            }
        )
        return context

    def _collect_entries(self, context: dict[str, object]) -> list[AuditLogEntry]:
        page_obj = context.get("page_obj")
        if hasattr(page_obj, "object_list"):
            return list(getattr(page_obj, "object_list"))
        entries = context.get(self.context_object_name)
        if isinstance(entries, list):
            return entries
        return list(entries or [])

    def _prepare_audit_map(
        self, entries: Iterable[AuditLogEntry]
    ) -> dict[int, Audit]:
        audit_ids = {
            int(audit_id)
            for audit_id in (
                getattr(entry, "related_audit_id", None) for entry in entries
            )
            if audit_id
        }
        if not audit_ids:
            return {}
        audits = (
            Audit.objects.filter(pk__in=audit_ids)
            .select_related("elevator", "elevator__building")
        )
        return {audit.pk: audit for audit in audits}

    def _build_entity_type_choices(self) -> list[tuple[str, str]]:
        choices = [("", _("Все объекты"))]
        for key, label in sorted(self.ENTITY_TYPE_LABELS.items(), key=lambda item: item[1]):
            choices.append((key, label))
        return choices

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.isoformat() if value else ""

    @staticmethod
    def _format_payload(payload: object) -> str:
        try:
            return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)
        except TypeError:
            return str(payload)

    def get_preserved_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        params.pop("export", None)
        non_empty = {key: value for key, value in params.items() if value}
        return urlencode(non_empty)

    def _export_csv(self, queryset: QuerySet[AuditLogEntry]) -> HttpResponse:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="audit-log.csv"'
        writer = csv.writer(response)
        writer.writerow(["created_at", "action", "entity", "user", "payload"])
        for entry in queryset.iterator():
            writer.writerow(
                [
                    timezone.localtime(entry.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                    entry.get_action_display(),
                    f"{entry.entity_type}:{entry.entity_id}",
                    getattr(entry.user, "get_username", lambda: "")(),
                    json.dumps(entry.payload, ensure_ascii=False, sort_keys=True),
                ]
            )
        return response


class OfflineSyncBatchListView(RoleRequiredMixin, ListView):
    """Monitoring page for offline synchronisation batches."""

    model = OfflineSyncBatch
    template_name = "audits/offline_batch_list.html"
    context_object_name = "batches"
    paginate_by = 25
    ordering = "-created_at"
    allowed_roles = (UserProfile.Roles.ADMIN,)

    start_param = "start"
    end_param = "end"
    status_param = "status"
    device_param = "device"

    _start_date: date | None = None
    _end_date: date | None = None
    _status_filter: str | None = None
    _device_filter: str | None = None

    def get(self, request, *args, **kwargs):  # type: ignore[override]
        if request.GET.get("export") == "csv":
            queryset = self.get_filtered_queryset()
            return self._export_csv(queryset)
        return super().get(request, *args, **kwargs)

    def get_queryset(self) -> QuerySet[OfflineSyncBatch]:  # type: ignore[override]
        return self.get_filtered_queryset()

    def get_filtered_queryset(self) -> QuerySet[OfflineSyncBatch]:
        queryset = super().get_queryset().select_related("user")
        queryset = self._filter_by_dates(queryset)
        queryset = self._filter_by_status(queryset)
        queryset = self._filter_by_device(queryset)
        return queryset.order_by(self.ordering)

    def _filter_by_dates(
        self, queryset: QuerySet[OfflineSyncBatch]
    ) -> QuerySet[OfflineSyncBatch]:
        start_date = self.get_start_date()
        if start_date:
            start_dt = timezone.make_aware(
                datetime.combine(start_date, time.min),
                timezone.get_current_timezone(),
            )
            queryset = queryset.filter(created_at__gte=start_dt)

        end_date = self.get_end_date()
        if end_date:
            exclusive_end = end_date + timedelta(days=1)
            end_dt = timezone.make_aware(
                datetime.combine(exclusive_end, time.min),
                timezone.get_current_timezone(),
            )
            queryset = queryset.filter(created_at__lt=end_dt)
        return queryset

    def _filter_by_status(
        self, queryset: QuerySet[OfflineSyncBatch]
    ) -> QuerySet[OfflineSyncBatch]:
        status_value = self.get_status_filter()
        valid_values = {value for value, _ in OfflineSyncBatch.Status.choices}
        if status_value and status_value in valid_values:
            queryset = queryset.filter(status=status_value)
        return queryset

    def _filter_by_device(
        self, queryset: QuerySet[OfflineSyncBatch]
    ) -> QuerySet[OfflineSyncBatch]:
        device_value = self.get_device_filter()
        if device_value:
            queryset = queryset.filter(device_id__icontains=device_value)
        return queryset

    def get_start_date(self) -> date | None:
        if self._start_date is not None:
            return self._start_date
        raw = self.request.GET.get(self.start_param, "").strip()
        self._start_date = parse_date(raw)
        return self._start_date

    def get_end_date(self) -> date | None:
        if self._end_date is not None:
            return self._end_date
        raw = self.request.GET.get(self.end_param, "").strip()
        self._end_date = parse_date(raw)
        return self._end_date

    def get_status_filter(self) -> str:
        if self._status_filter is not None:
            return self._status_filter
        self._status_filter = self.request.GET.get(self.status_param, "").strip()
        return self._status_filter

    def get_device_filter(self) -> str:
        if self._device_filter is not None:
            return self._device_filter
        self._device_filter = self.request.GET.get(self.device_param, "").strip()
        return self._device_filter

    def get_context_data(self, **kwargs: object) -> dict[str, object]:
        context = super().get_context_data(**kwargs)
        context.update(
            {
                "start_value": self._format_date(self.get_start_date()),
                "end_value": self._format_date(self.get_end_date()),
                "status_value": self.get_status_filter(),
                "status_choices": self._build_status_choices(),
                "device_value": self.get_device_filter(),
                "querystring": self.get_preserved_querystring(),
                "notification_choices": OfflineSyncBatch.NotificationStatus.choices,
                "status_summary": self._build_status_summary(),
                "start_param": self.start_param,
                "end_param": self.end_param,
                "status_param": self.status_param,
                "device_param": self.device_param,
            }
        )
        for batch in self._collect_batches(context):
            setattr(batch, "error_details_pretty", self._format_payload(batch.error_details))
        return context

    def _build_status_choices(self) -> list[tuple[str, str]]:
        choices = [("", _("Все статусы"))]
        choices.extend(OfflineSyncBatch.Status.choices)
        return choices

    def _build_status_summary(self) -> dict[str, int]:
        base_queryset = OfflineSyncBatch.objects.all()
        return {
            "total": base_queryset.count(),
            OfflineSyncBatch.Status.PENDING: base_queryset.filter(
                status=OfflineSyncBatch.Status.PENDING
            ).count(),
            OfflineSyncBatch.Status.APPLIED: base_queryset.filter(
                status=OfflineSyncBatch.Status.APPLIED
            ).count(),
            OfflineSyncBatch.Status.ERROR: base_queryset.filter(
                status=OfflineSyncBatch.Status.ERROR
            ).count(),
        }

    @staticmethod
    def _format_date(value: date | None) -> str:
        return value.isoformat() if value else ""

    @staticmethod
    def _format_payload(payload: object) -> str:
        try:
            return json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True)
        except TypeError:
            return str(payload)

    def _collect_batches(self, context: dict[str, object]) -> list[OfflineSyncBatch]:
        page_obj = context.get("page_obj")
        if hasattr(page_obj, "object_list"):
            return list(getattr(page_obj, "object_list"))
        batches = context.get(self.context_object_name)
        if isinstance(batches, list):
            return batches
        return list(batches or [])

    def get_preserved_querystring(self) -> str:
        params = self.request.GET.copy()
        params.pop("page", None)
        params.pop("export", None)
        non_empty = {key: value for key, value in params.items() if value}
        return urlencode(non_empty)

    def _export_csv(self, queryset: QuerySet[OfflineSyncBatch]) -> HttpResponse:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="offline-batches.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "created_at",
                "device_id",
                "status",
                "user",
                "response_status",
                "notification_status",
                "notified_at",
                "error_details",
            ]
        )
        for batch in queryset.iterator():
            writer.writerow(
                [
                    timezone.localtime(batch.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                    batch.device_id,
                    batch.get_status_display(),
                    getattr(batch.user, "get_username", lambda: "")(),
                    batch.response_status,
                    batch.get_error_notification_status_display(),
                    batch.error_notified_at.isoformat()
                    if batch.error_notified_at
                    else "",
                    json.dumps(batch.error_details, ensure_ascii=False, sort_keys=True),
                ]
            )
        return response

__all__ = [
    "AuditListView",
    "AttachmentDownloadView",
    "AuditCSVExportView",
    "AuditExcelExportView",
    "AuditLogEntryListView",
    "AuditPrintView",
    "OfflineChecklistView",
    "OfflineSyncBatchListView",
    "OfflineObjectInfoView",
]
