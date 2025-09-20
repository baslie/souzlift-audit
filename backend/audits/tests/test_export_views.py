"""Tests covering audit export endpoints (HTML, CSV, Excel)."""
from __future__ import annotations

import csv
import io

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone

from openpyxl import load_workbook

from accounts.models import UserProfile
from audits.models import Audit, AuditAttachment, AuditResponse
from catalog.models import (
    Building,
    ChecklistCategory,
    ChecklistQuestion,
    ChecklistSection,
    Elevator,
    ObjectInfoField,
)

from .test_attachment_access import ProtectedMediaTestCase, SMALL_GIF


class AuditExportViewTests(ProtectedMediaTestCase):
    """Ensure audit export views return data for authorized users only."""

    def setUp(self) -> None:
        super().setUp()
        UserModel = get_user_model()

        self.admin = UserModel.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="StrongPass123",
        )
        self.admin.profile.role = UserProfile.Roles.ADMIN
        self.admin.profile.save(update_fields=["role"])
        self.admin.profile.mark_password_changed()

        self.auditor = UserModel.objects.create_user(
            username="auditor",
            email="auditor@example.com",
            password="StrongPass123",
        )
        self.auditor.profile.role = UserProfile.Roles.AUDITOR
        self.auditor.profile.save(update_fields=["role"])
        self.auditor.profile.mark_password_changed()

        self.other_auditor = UserModel.objects.create_user(
            username="auditor2",
            email="auditor2@example.com",
            password="StrongPass123",
        )
        self.other_auditor.profile.role = UserProfile.Roles.AUDITOR
        self.other_auditor.profile.save(update_fields=["role"])
        self.other_auditor.profile.mark_password_changed()

        self.building = Building.objects.create(address="Тверская, 10", created_by=self.admin)
        self.elevator = Elevator.objects.create(
            building=self.building,
            identifier="EL-200",
            created_by=self.admin,
        )

        ObjectInfoField.objects.create(
            code="manager",
            label="Ответственный",
            field_type=ObjectInfoField.FieldType.TEXT,
            order=1,
        )
        ObjectInfoField.objects.create(
            code="capacity",
            label="Вместимость",
            field_type=ObjectInfoField.FieldType.NUMBER,
            order=2,
        )

        category = ChecklistCategory.objects.create(code="safety", name="Безопасность", order=1)
        section = ChecklistSection.objects.create(category=category, title="Основные проверки", order=1)
        self.score_question = ChecklistQuestion.objects.create(
            section=section,
            text="Состояние дверей",
            max_score=5,
            order=1,
        )
        self.text_question = ChecklistQuestion.objects.create(
            section=section,
            text="Дополнительные замечания",
            type=ChecklistQuestion.QuestionType.TEXT,
            order=2,
        )

        planned_date = timezone.localdate()
        self.audit = Audit.objects.create(
            elevator=self.elevator,
            created_by=self.auditor,
            planned_date=planned_date,
            object_info={
                "manager": "Иван Иванов",
                "capacity": 10,
            },
            total_score=4,
        )
        self.audit.start(actor=self.auditor)
        self.audit.submit(actor=self.auditor)

        self.score_response = AuditResponse.objects.create(
            audit=self.audit,
            question=self.score_question,
            score=4,
            comment="Требуется регулировка",
            is_flagged=True,
        )
        self.text_response = AuditResponse.objects.create(
            audit=self.audit,
            question=self.text_question,
            comment="Описание состояния площадки",
        )

        AuditAttachment.objects.create(
            response=self.score_response,
            file=SimpleUploadedFile("door.gif", SMALL_GIF, content_type="image/gif"),
            caption="Фото двери",
        )

    def test_print_view_renders_audit_details(self) -> None:
        """Printable report should display key audit information."""

        self.client.force_login(self.admin)
        url = reverse("audits:audit-export-print", args=[self.audit.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f"Аудит №{self.audit.pk}")
        self.assertContains(response, "Состояние дверей")
        self.assertContains(response, "Описание состояния площадки")
        self.assertContains(response, "Фото двери")

    def test_csv_export_contains_question_rows(self) -> None:
        """CSV export should include checklist data and attachment links."""

        self.client.force_login(self.admin)
        url = reverse("audits:audit-export-csv", args=[self.audit.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")

        content = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content), delimiter=";")
        rows = list(reader)

        header = [
            "Категория",
            "Раздел",
            "Вопрос",
            "Ответ",
            "Комментарий",
            "Пометка",
            "Вложения",
        ]
        self.assertIn(header, rows)
        header_index = rows.index(header)
        data_row = rows[header_index + 1]

        self.assertEqual(data_row[0], "Безопасность")
        self.assertEqual(data_row[2], "Состояние дверей")
        self.assertEqual(data_row[3], "4 из 5")
        self.assertIn("http://testserver", data_row[6])

    def test_excel_export_contains_question_rows(self) -> None:
        """Excel export should provide structured worksheet with answers."""

        self.client.force_login(self.admin)
        url = reverse("audits:audit-export-excel", args=[self.audit.pk])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(filename=io.BytesIO(response.content))
        sheet = workbook.active

        # Locate the first checklist row after the header.
        header_row = None
        for row in sheet.iter_rows(values_only=True):
            if row[:3] == ("Категория", "Раздел", "Вопрос"):
                header_row = row
                break
        self.assertIsNotNone(header_row)

        for row in sheet.iter_rows(values_only=True):
            if row[2] == "Состояние дверей":
                self.assertEqual(row[0], "Безопасность")
                self.assertEqual(row[3], "4 из 5")
                self.assertIn("http://testserver", row[6])
                break
        else:  # pragma: no cover - defensive
            self.fail("Checklist data row not found in Excel export")

    def test_auditor_can_export_own_audit(self) -> None:
        """Audit author should be able to download exports."""

        self.client.force_login(self.auditor)
        url = reverse("audits:audit-export-csv", args=[self.audit.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_other_auditor_cannot_access_exports(self) -> None:
        """Foreign auditors must not access someone else's audit export."""

        self.client.force_login(self.other_auditor)
        url = reverse("audits:audit-export-csv", args=[self.audit.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
